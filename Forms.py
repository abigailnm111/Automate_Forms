#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 20:43:31 2020

@author: media
"""


#import docx
import openpyxl
import re
import os
from math import ceil

import write_pre_offer as offer

#############################################################
#IMPORT DATA FROM ASSIGNMENT SPREADSHEET
location= os.path.dirname(os.path.abspath(__file__))
course_list=os.path.join(location,'test_courses.txt')
def get_courseloads(file):
    
    filename=os.path.join(location, file)
    full_assign= openpyxl.load_workbook(filename)
    year_assign= full_assign.active
    return year_assign


#assign cells to variables by line
#reads courseload spreadsheet to create list of data for each employee
class Faculty:
    def __init__(self, year_assign, row, cont_range, pre_range):
                self.action_type=["Reappointment"] 
                self.warnings=[]
                
                row=str(row+1)
                i="A"+row
                self.last_name= re.findall('([A-Za-z]*),',year_assign[i].value)[0]
                self.first_name=re.findall(',\s([A-Za-z]*)', year_assign[i].value)[0]
                i= "B"+row                      
                self.course_total= re.findall('([1-9])\s', year_assign[i].value)[0]
                self.percentage= re.search('\d*%', year_assign[i].value).group()
                i="C"+row
                if year_assign[i].value != None:
                    self.F_courses= re.findall('[1-3]\s[A-Za-z]{2,4}\s\d{1,3}[a-zA-Z]*', year_assign[i].value)
                    cr= re.findall('[1-3]\s[cC]ourse\s[rR]elease', year_assign[i].value)
                    self.F_courses=self.get_course_names(self.F_courses)
                    if cr != []:
                        self.F_courses.append(cr[0])
                    
                else:
                    self.F_courses=[]
                i="D"+row
                if year_assign[i].value != None:
                    self.W_courses= re.findall('[1-3]\s[A-Za-z]{2,4}\s\d{1,3}[a-zA-Z]*', year_assign[i].value)
                    cr= re.findall('[1-3]\s[cC]ourse\s[rR]elease', year_assign[i].value)
                    self.W_courses=self.get_course_names(self.W_courses)
                    if cr != []:
                        self.W_courses.append(cr[0])
                   
                else:
                    self.W_courses=[]
                i="E"+row
                if year_assign[i].value != None:
                    self.S_courses=re.findall('[1-3]\s[A-Za-z]{2,4}\s\d{1,3}[a-zA-Z]*', year_assign[i].value)
                    cr= re.findall('[1-3]\s[cC]ourse\s[rR]elease', year_assign[i].value)
                    
                    self.S_courses=self.get_course_names(self.S_courses)
                    print(self.S_courses)
                    if cr != []:
                        self.S_courses.append(cr[0])
                    
                else:
                    self.S_courses=[]
                    
 #counts the number of active courses for the year 
    def get_quarters(self):  
        self.quarters=0
        if self.F_courses != []:
            
            self.quarters+=1
        if self.W_courses != []:
            self.quarters+=1
        if self.S_courses != []:
            self.quarters+=1
        return self.quarters
    
 #get effective and end dates based on total number of quarters and specific quarters active   
    def get_start_end_date(self, year):
            year=int(year)             
            if self.quarters== 3:
                self.start= ["7/1/"+ str(year)]
                self.end=["6/30/"+str(year+1)]
            if self.quarters==2:
                if self.F_courses != []:
                    if self.S_courses !=[]:
                        self.start= ["10/1/"+str(year), "4/1/"+ str(year+1)]
                        self.end=["12/31/"+str(year+1), "6/30/"+str(year+1)]
                    else:
                        self.start= ["10/1/"+ str(year)]
                        self.end= ["3/30/"+str(year+1)]
                elif self.W_courses != []:
                    self.start= ["1/1/"+ str(year+1)]
                    self.end= ["6/30/"+str(year+1)]
            if self.quarters==1:
                if self.F_courses !=[]:
                    self.start=["10/1/"+str(year)]
                    self.end= ["12/31/"+str(year)]
                elif self.W_courses!=[]:
                    self.start= ["1/1/"+str(year+1)]
                    self.end= ["3/30/"+str(year+1)]
                else:
                    self.start=["4/1/"+str(year+1)]
                    self.end=["6/30/"+str(year+1)]
            return self.start, self.end
    
   
            
    
#job code dependent on number of quarters for that year and rank
#pre six = 1630,1632
#Continuing= 1631, 1633
#Senior Continuing= 1641, 1643
    def get_job_code(self):
        if self.status==["pre"]: 
            if self.quarters==3:
                self.job_code= [1630]
                self.mo_paid_over= 12
            else:
                self.job_code= [1632]
                self.mo_paid_over= 9
            self.title= ["Lecturer"]
            
        if self.status== ["cont"]:
            if self.quarters== 3:
                self.job_code= [1631]
                self.mo_paid_over= 12
            else:
                self.job_code= [1633]
                self.mo_paid_over= 9
            self.title= ["Continuing Lecturer"]
            
        if self.status== ['sr']:
            if self.quarters== 3:
                self.job_code= [1641]
                self.mo_paid_over= 12
            else:
                self.job_code= [1643]
                self.mo_paid_over= 9
            self.title= ["Senior Continuing Lecturer"]
        return self.job_code, self.title
    
    def get_last_value_in_history(self,sheet, base_salary, empty_row, increase):
        i=0
        for column in ["H","J", "O"]:
            
            list_of_values=[]
            for row in sheet[column+"2":(column+str(empty_row-1))]:
                for cell in row:
                        list_of_values.append(cell.value)
                        list_of_values=[q for q in list_of_values if q != None]              
            list_of_values.sort(reverse=True)
            
            if list_of_values!=[]:
                q=list_of_values[0]
            else:
                q=0
            if i==0:
                self.HRquarters=q
                if self.HRquarters>=18:
                    self.status= ['cont']
                     
                else:
                    self.status=['pre']
                
            elif (i==1) and (self.status== ["cont"]):
                if (1641 in list_of_values)or (1643 in list_of_values):
                    self.status=["sr"]
                    
            elif i==2:
                self.get_job_code()
                if q==0:
                    self.annual= [base_salary]
                    self.warnings.append(self.last_name+": base salary used")
                else:
                    self.annual=[ceil(q*(1+increase/100))]
                self.monthly= [self.annual[0]/self.mo_paid_over]
            
            
            i+=1
        return self.HRquarters, self.annual, self.monthly
       
    def pre6_milestone_check(self, AY, q_check, action):
            i=1
            if self.F_courses!=[] and (self.HRquarters+i)<=q_check:
                if (self.HRquarters+i)==q_check:
                    self.action_type.insert(0,action)
                    self.annual= ["ATTENTION"]
                    self.monthly= ["ATTENTION"]                  
                    return 
                else:
                    i+=1
            if self.W_courses!=[]and (self.HRquarters+i)==q_check:
                
                if self.quarters==3:
                    self.start.append("11/1/"+AY)
                    self.end.insert(0, "10/31/"+AY)
                    if q_check==19:
                        self.job_code.append("1631")
                        self.title.append("Continuing Lecturer")
                else:
                    if ("1/1/"+str(int(AY)+1)) not in self.start:
                        self.start.append("1/1/"+str(int(AY)+1))
                        self.end.insert(0, "12/31/"+AY)
                        if q_check==19:
                            self.job_code.append("1633")
                            self.title.append("Continuing Lecturer")
            else:
                
                if self.quarters==3:
                    self.start.append("3/1"+str(int(AY)+1))
                    self.end.insert(0, "2/28/"+str(int(AY)+1)) 
                    self.warnings.append(self.last_name+": End date set to 2/28.Check leap year end date")
                    if q_check==19:
                        self.job_code.append("1631")
                        self.title.append("Continuing Lecturer")
                else:
                    
                            
                    if ("4/1/"+str(int(AY)+1)) not in self.start:
                        
                        self.start.append("4/1/"+str(int(AY)+1))
                        self.end.insert(0, "3/31/"+str(int(AY)+1)) 
                        if q_check==19:
                            self.title.append("Continuing Lecturer")
                            self.job_code.append("1633")
            if len(self.start)>1:
                self.action_type.append(action)
                self.annual.append("ATTENTION")
                self.monthly.append( "ATTENTION")
            else:
                self.action_type.insert(0,action)
                self.annual= ["ATTENTION"]
                self.monthly= ["ATTENTION"]
            return self.start, self.end
        
    def get_course_names(self, quarter):      
        course_names=open(course_list,'r')
        
        course_title_list=[]
        i=1
        for course in quarter:
            title= course.split(' ')
            
            for c in course_names:
                print(title[2])
                course_title=re.search("%s[.].+"%title[2],c)
                if course_title!= None:
                    if i==len(quarter):
                        course_title_list.append(title[0]+" "+course_title.group()+", ")
                    else:
                        course_title_list.append(title[0]+" "+course_title.group())
                i+=1
        return course_title_list    
   

def get_rank_ranges(year_assign):
    for row in year_assign.rows:
        if row[0].value=="CONTINUING":
            
            cont_start_row=row[0].row
            cont_start_row= int(cont_start_row)
            
        if row[0].value== "PRE-SIX":
            pre_start_row=row[0].row
            pre_start_row= int(pre_start_row)
    return cont_start_row, pre_start_row
        


    
#########################################################
#HISTORY RECORD

####Get quarter count from prior History Record


def update_history_record(lecturer, a_sheet,AY, base_salary, HR, increase):
    #adds name to History Record Header if Template used
    if HR== "new":
        a_sheet.HeaderFooter.scaleWithDoc= True
        a_sheet.HeaderFooter.evenHeader.left.text=lecturer.last_name+' '+ lecturer.first_name
        a_sheet.HeaderFooter.oddHeader.left.text=lecturer.last_name+" "+lecturer.first_name
        
        
    for cell in a_sheet["B"]:
        if cell.value is None:
            
            #sets Academic Year
            a_sheet.cell(row=cell.row, column=1).value= AY+"-"+str(int(AY)+1)
            lecturer.get_quarters(), lecturer.get_start_end_date(AY)
            #gets pre-exisiting History Record information unless template is used
          
            if HR!="new":
                if increase>0:
                    lecturer.action_type[0]= ("Reappointment Range Adjustment "+str(increase)+"%")
                lecturer.get_last_value_in_history(a_sheet, base_salary, cell.row, increase)
                total_quarters= lecturer.HRquarters+lecturer.quarters
                
                
                ##checks for 9th quarter to alert user- no affect on documents
                if 9 in range(lecturer.HRquarters+1,total_quarters+1):
                    lecturer.warnings.append(lecturer.last_name+":9th quarter warning- check for 9th quarter mentoring meeting")
                
                #checks for Pre 6 miles stones that require additional action. Does not input salary into documents and alerts User.
                if 10 in range(lecturer.HRquarters+1,total_quarters+1):
                    lecturer.warnings.append(lecturer.last_name+":10th quarter warning- check for 10th quarter increase")
                    lecturer.pre6_milestone_check(AY, 10, "10th Quarter Increase")
                    
                if 19 in range(lecturer.HRquarters+1,total_quarters+1):
                    lecturer.warnings.append(lecturer.last_name+":19th quarter warning- check for Continuing Appointment")
                    lecturer.pre6_milestone_check(AY, 19, "Initial Continuing Appointment")    
            else:
                lecturer.HRquarters=0
                lecturer.status=["pre"]
                lecturer.get_job_code()
                lecturer.annual= [base_salary]
                lecturer.monthly= [lecturer.annual[0]/lecturer.mo_paid_over]
                lecturer.warnings.append(lecturer.last_name+": base salary used")
                
            
                    
            i=0
            for start in lecturer.start:
                
                
                ##checks for 9th quarter to alert user- no affect on documents
                
                
                a_sheet.cell(row=cell.row+i, column=2).value=lecturer.start[i]
                a_sheet.cell(row=cell.row+i,column=3).value=lecturer.end[i]
                #accounts for a F/S appointment with a break 
                if len(lecturer.start)>1:
                   
                   if ("10/1" in lecturer.start[i]) or( "7/1" in lecturer.start[i]):  
                       a_sheet.cell(row=cell.row+i,column=5).value="x"
                       lecturer.HRquarters+=1
                       if (lecturer.W_courses!=[]) and ("1/1" not in lecturer.start[i+1]):
                           a_sheet.cell(row=cell.row+i,column=6).value= "x"
                           lecturer.HRquarters+=1
                   if ("11/1/" in lecturer.start[i]) or ("1/1/" in lecturer.start[i]):
                       a_sheet.cell(row=cell.row+i,column=6).value= "x"
                       lecturer.HRquarters+=1
                       if (lecturer.S_courses !=[]) and (i== len(lecturer.start)-1):
                           a_sheet.cell(row=cell.row+i,column=7).value= "x"
                           lecturer.HRquarters+=1                        
                   if ("4/1" in lecturer.start[i] )or ("3/1" in lecturer.start[i]):
                       a_sheet.cell(row=cell.row+i,column=7).value= "x"
                       lecturer.HRquarters+=1
                   a_sheet.cell(row=cell.row+i,column=8).value=lecturer.HRquarters
                   a_sheet.cell(row=cell.row+i,column=9).value= (lecturer.HRquarters)/3
                #all other single quarter or continuious appointments
               
                else:
                    if lecturer.F_courses != []:
                        a_sheet.cell(row=cell.row+i,column=5).value="x"
                
                    if lecturer.W_courses != []:
                        a_sheet.cell(row=cell.row+i,column=6).value="x"
                    
                    if lecturer.S_courses != []:
                        a_sheet.cell(row=cell.row+i,column=7).value= "x"
                    a_sheet.cell(row=cell.row+i,column=8).value=lecturer.HRquarters+lecturer.quarters
                    a_sheet.cell(row=cell.row+i,column=9).value= (lecturer.HRquarters+lecturer.quarters)/3
                
                
               
                if len(lecturer.action_type)==1:
                    
                    a_sheet.cell(row=cell.row+i,column=12).value=lecturer.action_type[0]
                    a_sheet.cell(row=cell.row+i,column=14).value=lecturer.monthly[0]
                    a_sheet.cell(row=cell.row+i,column=15).value=lecturer.annual[0]
                else:
                    
                        
                    a_sheet.cell(row=cell.row+i,column=12).value=lecturer.action_type[i]
                    a_sheet.cell(row=cell.row+i,column=14).value=lecturer.monthly[i]
                    a_sheet.cell(row=cell.row+i,column=15).value=lecturer.annual[i]
                    
                if len(lecturer.job_code)==1:
                    a_sheet.cell(row=cell.row+i,column=10).value=lecturer.job_code[0]
                    a_sheet.cell(row=cell.row+i,column=11).value=lecturer.title[0]
                else:
                    a_sheet.cell(row=cell.row+i,column=10).value=lecturer.job_code[i]
                    a_sheet.cell(row=cell.row+i,column=11).value=lecturer.title[i]
                a_sheet.cell(row=cell.row+i,column=13).value=lecturer.percentage

                i+=1
            break
        
  ################################################     
#read document and replace key words

def get_first_empty_cell(ws,column):
    for cell in ws[column]:
        if cell.value is None:
            
            return cell.row
#write document

def main():
    AY= input("enter first year of academic year Ex. '20XX'")
    file=input("enter courseload file name: ex: courseload AY.xslx")
    base_salary= int(input("a base annual salary will be inputed for faculty without a salary history. What base salary do you want to be entered?(Do not include a '$' or comma)Base salary should include any Range Adjustments."))
    RA= input("Is there a Range adjustment effective 7/1/"+AY+"? ('Y' or' N')")
    if (RA== "Y") or (RA== "y"):
        increase= int(input("enter the percentage amount (number only Ex: 'x'%)"))
    else:
        increase=0
    year_assign=get_courseloads(file)
    all_lec=[]
    c_range, p_range=get_rank_ranges(year_assign)
    assignment_len= get_first_empty_cell(year_assign,"A")
    
#iterates through class to read courseload spreadsheet but skips Pre-Six header row
    for row in range (2, assignment_len-1):
        if row!= (p_range-1):
            
            all_lec.append(Faculty(year_assign, row, c_range, p_range))
#looks for current History Record to Copy and Creates a new one from template if does not exist        
    for faculty in all_lec: 
        
        HR_file= os.path.join(location, faculty.last_name+"."+faculty.first_name+".xlsx" )
        HR_template= os.path.join(location, "2_History_Record_template.xlsx" )
        try:
            wp=openpyxl.load_workbook(HR_file)
            dest_filename= os.path.join(location, faculty.last_name+"." +faculty.first_name+".xlsx")
            HR=""
        except:
            wp=openpyxl.load_workbook(HR_template)
            dest_filename= os.path.join(location, "2_History_Record_template.xlsx")
            HR="new"
            
            
        wp=openpyxl.load_workbook(dest_filename)
        a_sheet=wp.active
        update_history_record(faculty, a_sheet, AY, base_salary, HR, increase)
        offer.write_pre_offer_letter(faculty)
        
        for warning in faculty.warnings:
            print (warning)
        wp.save(os.path.join(location, AY+faculty.last_name+"." +faculty.first_name+".xlsx"))
main()